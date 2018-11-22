#coding=utf-8
import xlwt
import xlrd
import types
import openpyxl
import re
import pdb
from openpyxl.reader.excel import load_workbook
import argparse
from xlsxwriter.workbook import Workbook

class ExcelTemplete:

	def __init__(self,ptn_id):
		if 1 == ptn_id:
			self.get_premium_discount_pattern();
		else:
			self.get_full_pattern();
	
	def get_premium_discount_pattern(self):
                self.head_row = 0;
                self.content_row = 1;
		self.head_list = [
				[u'批号' , [0,0]],
				[u'11' , [0,1]],
				[u'21' , [0,2]],
				[u'31' , [0,3]],
				[u'41' , [0,4]],
				[u'51' , [0,5]],
				[u'12' , [0,6]],
				[u'22' , [0,7]],
				[u'32' , [0,8]],
				[u'13' , [0,9]],
				[u'23' , [0,10]],
				[u'33' , [0,11]],
				[u'14' , [0,12]],
				[u'24' , [0,13]],
				[u'平均长度' , [0,14]],
				[u'平均马值' , [0,15]],
				[u'平均断裂比强度' , [0,16]],
				[u'平均长度整齐度' , [0,17]],
				[u'P1' , [0,18]],
				[u'P2' , [0,19]],
				[u'P3' , [0,20]],
				];
		self.dic_pos = {
				'production_code':0,
				'colour_w1':1,
				'colour_w2':2,
				'colour_w3':3,
				'colour_w4':4,
				'colour_w5':5,
				'colour_l1':6,
				'colour_l2':7,
				'colour_l3':8,
				'colour_ly1':9,
				'colour_ly2':10,
				'colour_ly3':11,
				'colour_y1':12,
				'colour_y2':13,
				'avg_length':14,
				'avg_micronaire':15,
				'strength':16,
				'avg_evenness':17,
				'ginning_p1':18,
				'ginning_p2':19,
				'ginning_p3':20,
				};

	def get_full_pattern(self):
                self.head_row = 0;
                self.content_row = 2;
		self.head_list = [
				[u'批号' , [0,1,0,0]],
				[u'颜色级' , [0,0,1,13]],
				[u'11' , [1,1]],
				[u'21' , [1,2]],
				[u'31' , [1,3]],
				[u'41' , [1,4]],
				[u'51' , [1,5]],
				[u'12' , [1,6]],
				[u'22' , [1,7]],
				[u'32' , [1,8]],
				[u'13' , [1,9]],
				[u'23' , [1,10]],
				[u'33' , [1,11]],
				[u'14' , [1,12]],
				[u'24' , [1,13]],
				[u'平均长度' , [0,1,14,14]],
				[u'长度分布' , [0,0,15,22]],
				[u'32mm' , [1,15]],
				[u'31mm' , [1,16]],
				[u'30mm' , [1,17]],
				[u'29mm' , [1,18]],
				[u'28mm' , [1,19]],
				[u'27mm' , [1,20]],
				[u'26mm' , [1,21]],
				[u'25mm' , [1,22]],
				[u'马克隆平均值' , [0,1,23,23]],
				[u'马值分布' , [0,0,24,28]],
				[u'C1' , [1,24]],
				[u'B1' , [1,25]],
				[u'A' , [1,26]],
				[u'B2' , [1,27]],
				[u'C2' , [1,28]],
				[u'断裂比强度' , [0,0,29,31]],
				[u'平均值' , [1,29]],
				[u'最大值' , [1,30]],
				[u'最小值' , [1,31]],
				[u'长度整齐度' , [0,0,32,34]],
				[u'平均值' , [1,32]],
				[u'最大值' , [1,33]],
				[u'最小值' , [1,34]],
				[u'轧工质量' , [0,0,35,37]],
				[u'P1' , [1,35]],
				[u'P2' , [1,36]],
				[u'P3' , [1,37]],
				[u'合计包数' , [0,1,38,38]],
				[u'合计毛重' , [0,1,39,39]],
				[u'合计皮重' , [0,1,40,40]],
				[u'合计净重' , [0,1,41,41]],
				[u'合计公重' , [0,1,42,42]],
				[u'平均回潮' , [0,1,43,43]],
				[u'平均含杂' , [0,1,44,44]],
				[u'产地' , [0,1,45,45]],
				[u'加工类型' , [0,1,46,46]],
				[u'加工企业' , [0,1,47,47]],
				[u'仓库' , [0,1,48,48]],
				];
		self.dic_pos = {
				'production_code':0,
				'colour_w1':1,
				'colour_w2':2,
				'colour_w3':3,
				'colour_w4':4,
				'colour_w5':5,
				'colour_l1':6,
				'colour_l2':7,
				'colour_l3':8,
				'colour_ly1':9,
				'colour_ly2':10,
				'colour_ly3':11,
				'colour_y1':12,
				'colour_y2':13,
				'avg_length':14,
				'length_32':15,
				'length_31':16,
				'length_30':17,
				'length_29':18,
				'length_28':19,
				'length_27':20,
				'length_26':21,
				'length_25':22,
				'avg_micronaire':23,
				'micronaire_c1':24,
				'micronaire_b1':25,
				'micronaire_a':26,
				'micronaire_b2':27,
				'micronaire_c2':28,
				'strength':29,
				'strength_max':30,
				'strength_min':31,
				'avg_evenness':32,
				'evenness_max':33,
				'evenness_min':34,
				'ginning_p1':35,
				'ginning_p2':36,
				'ginning_p3':37,
				'package_num':38,
				'weight_gross':39,
				'weight_tare':40,
				'weight_net':41,
				'weight_conditoned':42,
				'huichao':43,
				'miscellaneous':44,
				'jiagongleixing':46,
				'factory':47,
				'warehouse':48,
				'production_area':45,
				};


class FileExcel_XLSX_Writer:

	def __init__(self):
		self.ptn_batch_num = re.compile(r'^[0-9]+$');

	def write_header(self, table, template):
                start_row = template.head_row;
		head_list = template.head_list;
		for item in head_list:
			if len(item) != 2:
				continue;
			#label = item[0].encode('utf-8');
			label = item[0];
			pos = item[1];
			if 2 == len(pos):
				table.write(start_row+pos[0], pos[1], label);
			elif 4 == len(pos):
				table.merge_range( first_row= start_row+pos[0], 
					first_col= pos[2], 
					last_row= start_row + pos[1], 
					last_col= pos[3], data=label);
			else:
				continue;

	def write_content_iter(self, table, template, asin_iter, **kw):
		row_num = template.content_row;
		dic_pos = template.dic_pos;
		for info in asin_iter(num_year=kw.get('year',18)):
			if info is None or 'production_code' not in info:
				continue;
			info['production_code'] = str(info['production_code']);
			for k in info:
				if isinstance( info[k], bytearray):
					info[k] = info[k].decode(encoding="utf-8", errors="strict");
			row = [ info[k] for k in [item[0] for item in sorted(dic_pos.items(), key=lambda x:x[1])]];
			table.write_row(row_num, 0, row);
			row_num += 1;

	def save_file_iter(self, path, template, asin_iter, **kw):
		workbook = Workbook(path);
		worksheet = workbook.add_worksheet();
		self.write_header(worksheet, template);
		self.write_content_iter(worksheet, template, asin_iter, **kw);
		workbook.close();

class FileExcel_XLSX:

	def __init__(self):
		self.ptn_batch_num = re.compile(r'^[0-9]+$');
	
	def writeHead(self, table, start_row, template):
                start_row = template.head_row;
		head_list = template.head_list;
		for item in head_list:
			if len(item) != 2:
				continue;
			#label = item[0].encode('utf-8');
			label = item[0];
			pos = item[1];
			if 2 == len(pos):
				table.cell(row=start_row+pos[0]+1,column=pos[1]+1, value=label);
			elif 4 == len(pos):
				table.merge_cells(start_row=start_row+pos[0]+1,
					start_column=pos[2]+1,
					end_row=start_row + pos[1]+1,
					end_column=pos[3]+1);
				table.cell(row=start_row+pos[0]+1,column=pos[2]+1, value=label);
			else:
				continue;

	def write_content_iter(self, table, template, asin_iter, **kw):
		row_num = template.content_row;
		dic_pos = template.dic_pos;
		for info in asin_iter(num_year=kw.get('year',18)):
			if info is None or 'production_code' not in info:
				continue;
			for seg in info.keys():
				if seg not in dic_pos:
					continue;
				col = dic_pos[seg];
				val = info[seg];
				if isinstance( val, bytearray):
					val = val.decode(encoding="utf-8", errors="strict");
				table.cell(row=row_num+1,column=col+1,value=val);
			row_num += 1;
				
			
	def writeContent(self, table, start_row, dic_val, template, order_list):
                start_row = template.content_row;
		dic_pos = template.dic_pos;
		row_num = start_row-1;
		for idx in range(len(order_list)):
			production_code = order_list[idx];
			if None == production_code or production_code not in dic_val:
				continue;
			row_num += 1;
			if 'production_code' in dic_pos:
				col = dic_pos['production_code'];
				table.cell(row=row_num+1,column=col+1,value=u"%d"%(production_code));
			info = dic_val[production_code];
			if None == info:
				continue;
			for seg in info.keys():
				if seg not in dic_pos or 'production_code'==seg:
					continue;
				col = dic_pos[seg];
				val = info[seg];
				if isinstance( val, bytearray):
					val = val.decode(encoding="utf-8", errors="strict");
				table.cell(row=row_num+1,column=col+1,value=val);

	def SaveFile(self, path, dic_val, template, order_list):
		wb=openpyxl.Workbook();
		table = wb.active;
		self.writeHead(table, 0, template);
		self.writeContent(table, 2, dic_val, template, order_list);
		wb.save(path);


	def getBatchNumListFromExcel(self,filePath, columnId):
		try:
			wb=load_workbook(filePath);
		except Exception,e:
			raise e

class FileExcel_XLS:
	def __init__(self):
		self.ptn_batch_num = re.compile(r'^[0-9]+$');
		self.style_head = xlwt.XFStyle();
		self.style_content = xlwt.XFStyle();
		
	
	def writeHead(self, table, start_row, template):
                start_row = template.head_row;
		head_list = template.head_list;
		for item in head_list:
			if len(item) != 2:
				continue;
			#label = item[0].encode('utf-8');
			label = item[0];
			pos = item[1];
			if 2 == len(pos):
				table.write(start_row+pos[0], pos[1], label);
			elif 4 == len(pos):
				table.write_merge(start_row+pos[0],start_row + pos[1],pos[2],pos[3],label);
			else:
				continue;

	def writeContent(self, table, start_row, dic_val, template, order_list):
                start_row = template.content_row;
		dic_pos = template.dic_pos;
		row_num = start_row-1;
		for idx in range(len(order_list)):
			production_code = order_list[idx];
			if None == production_code or production_code not in dic_val:
				continue;
			row_num += 1;
			if 'production_code' in dic_pos:
				col = dic_pos['production_code'];
				table.write(row_num, col, u"%d"%(production_code));
			info = dic_val[production_code];
			if None == info:
				continue;
			for seg in info.keys():
				if seg not in dic_pos or seg == "production_code":
					continue;
				col = dic_pos[seg];
				val = info[seg];
				if isinstance( val, bytearray):
					val = val.decode(encoding="utf-8", errors="strict");
				table.write(row_num,col,val);

	def SaveFile(self, path, dic_val, template, order_list):
		wb = xlwt.Workbook(encoding='utf-8');
		table = wb.add_sheet(u'sheet1',cell_overwrite_ok=True);
		self.writeHead(table,0,template);
		self.writeContent(table,2,dic_val,template, order_list);
		wb.save(path);

	def getBatchNumListFromExcel(self,filePath, columnId):
		try:
			ret_list = [];
			data = xlrd.open_workbook(filePath);
			table = data.sheet_by_index(0);
			for i in range(table.nrows):
				val = table.cell(i, columnId).value;
				if type(val) is types.IntType:
					ret_list.append(str(val));
				elif type(val) is types.FloatType:
					ret_list.append(str("%.0f"%(val)));
				elif type(val) is types.StringType:
					if re.match(self.ptn_batch_num,val):
						ret_list.append(val);
				elif type(val) is types.UnicodeType:
					val = val.encode('utf-8');
					if re.match(self.ptn_batch_num,val):
						ret_list.append(val);
			return ret_list;
		except Exception,e:
			raise e



class FileManager:
	def __init__(self):
		self.file_excel_xls = FileExcel_XLS();
		self.file_excel_xlsx = FileExcel_XLSX();

	def ReadBatchNumList(self,file_path,column_id, file_type='excel'):
		if file_type == 'excel':
			return self.file_excel_xls.getBatchNumListFromExcel(file_path, column_id);

	def save(self, path, dic_val, template_id, order_list, file_type='excel'):
		template = ExcelTemplete(template_id);
		if file_type == 'excel':
			if path.endswith('xlsx'):
				self.file_excel_xlsx.SaveFile(path, dic_val, template, order_list);
			else:
				self.file_excel_xls.SaveFile(path, dic_val, template, order_list);

def write_file(year, output, template, **kw):
	from db_orm import qry_asins_by_year_iter;
	template = ExcelTemplete(template);
	file_excel_xlsx = FileExcel_XLSX_Writer();
	file_excel_xlsx.save_file_iter(output, template, qry_asins_by_year_iter, year=year);

def write():
	from db_orm import qry_asins_by_year; 
	file_mgr = FileManager();
	dic_info = qry_asins_by_year(18);
	#file_mgr.save('/home/zhangjn/cotton/apache-tomcat-7.0.88/webapps/futures-web/2018.xls', dic_info, 0, dic_info.keys());
	file_mgr.save('/home/zhangjn/cotton/apache-tomcat-7.0.88/webapps/futures-web/2018.xlsx', dic_info, 0, dic_info.keys());

if __name__ == '__main__':
	parser = argparse.ArgumentParser();
	parser.add_argument('--year',  type=int, default=18, required=False, help='input files name');
	parser.add_argument('--template',  type=int, default=0, required=False, help='input files name');
	parser.add_argument('--output', default='/home/zhangjn/cotton/apache-tomcat-7.0.88/webapps/futures-web/2018.xlsx', required=False, help='input files name');
	args = parser.parse_args();
	write_file(year=args.year, output=args.output, template=args.template);
