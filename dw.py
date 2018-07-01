#coding=utf-8
import requests
import json
from lxml import etree
import re
import multiprocessing
import time
import xlwt
import xlrd
import types
import openpyxl
from openpyxl.reader.excel import load_workbook
#import pdb
from db_tool import DataBase


time_per_req = 6;
time_in_req = 1;

class WebProcessor_Base:
	def __init__(self, url_home, explorer_type='firefox'):
		self.url_home = url_home;
		self.headers = {
				'Host': 'blog.csdn.net',
				'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:43.0) Gecko/20100101 Firefox/43.0',
				'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
				'Accept-Language': 'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
				'Accept-Encoding': 'gzip, deflate',
				'Referer': self.url_home,
				'Connection': 'keep-alive',
				'Cache-Control': 'max-age=0',
				};

	def normaliza_fields(self, batch_num, src_dic):
		if None == src_dic or not isinstance(src_dic, dict):
				return None;
		for key in src_dic:
			if src_dic[key] == '——':
				src_dic[key] = '';

		src_dic['batch_num'] = batch_num;
		if batch_num.startswith('65'):
			src_dic['chandi'] = '新疆地方';
		elif batch_num.startswith('66'):
			src_dic['chandi'] = '新疆兵团';
		else:
			if 'chandi' not in src_dic or len(src_dic['chandi'])<1:
				src_dic['chandi'] = '内地';
		
		if 'jiagongleixing' in src_dic:
			if src_dic['jiagongleixing'] in ['锯齿细绒棉']:
				src_dic['jiagongleixing'] = '手摘棉';
			elif src_dic['jiagongleixing'] in ['锯齿机采棉']:
				src_dic['jiagongleixing'] = '机采棉';

		for k in ['weight_gross','weight_tare','weight_net','weight_conditoned']:
			if k in src_dic:
				if src_dic[k].endswith('t') or src_dic[k].endswith('T'):
					src_dic[k] = src_dic[k][0:-1]
				elif src_dic[k].endswith('kg') or src_dic[k].endswith('KG'):
					src_dic[k] = src_dic[k][0:-2]
		return src_dic;

class WebProcessor_MZ(WebProcessor_Base):
	url = 'http://mianzhuang.com';
	url_detail_prefix = 'http://mianzhuang.com/ecommerce/control/quality?batchNo='
	ptn_preparation = re.compile('.*P1：(?P<p1>[\.\d]*)%[\s\S]*P2：(?P<p2>[\.\d]*)%[\s\S]*P3：(?P<p3>[\.\d]*)%.*')
	xpath_dic = {
			'color_11':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[8]/td[3]",
			'color_21':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[9]/td[2]",
			'color_31':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[10]/td[2]",
			'color_41':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[11]/td[2]",
			'color_51':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[12]/td[2]",
			'color_12':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[13]/td[2]",
			'color_22':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[14]/td[2]",
			'color_32':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[15]/td[2]",
			'color_13':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[16]/td[2]",
			'color_23':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[17]/td[2]",
			'color_33':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[18]/td[2]",
			'color_14':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[19]/td[2]",
			'color_24':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[20]/td[2]",
			'length_avg':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[7]/td[5]",
			'length_32':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[8]/td[6]",
			'length_31':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[9]/td[4]",
			'length_30':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[10]/td[4]",
			'length_29':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[11]/td[4]",
			'length_28':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[12]/td[4]",
			'length_27':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[13]/td[4]",
			'length_26':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[14]/td[4]",
			'length_25':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[15]/td[4]",
			'micronaire_avg':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[17]/td[4]",
			'micronaire_c1':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[18]/td[5]",
			'micronaire_b1':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[19]/td[4]",
			'micronaire_a':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[20]/td[4]",
			'micronaire_b2':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[21]/td[5]",
			'micronaire_c2':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[22]/td[4]",
			'breaking_tenacity_avg':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[25]/td[5]",
			'breaking_tenacity_max':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[24]/td[5]",
			'breaking_tenacity_min':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[23]/td[5]",
			'length_uniformty_avg':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[23]/td[2]",
			'length_uniformty_max':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[22]/td[2]",
			'length_uniformty_min':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[21]/td[3]",
			'preparation':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_search_sorting']/div[@id='mz_quality_table']/div[@id='report']/table[@class='table table-summary quality_table']/tr[5]/td[@class='txt-left'][2]",
			'package_num':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='packageSize']",
			'weight_gross':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='weightGross']",
			'weight_tare':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='weightTare']",
			'weight_net':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='weightNetPackage']",
			'weight_conditoned':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='weightPublic']",
			'huichao_avg':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='percentWet']",
			'hanza_avg':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='percentDust']",
			'jiagongleixing':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tr/td[@class='mz_quality_line02']/table[@id='quality_detail02']/tr[2]/td[@id='cottonType']",
			'jiagongqiye':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tbody[1]/tr[@class='mz_Watching_cont_tr']/td[5]",
			'cangku':"/html/body/div[@class='whole-container']/div[@id='mz_main']/div[@id='mz_quality']/table[@id='quality_detail01']/tbody[1]/tr[@class='mz_Watching_cont_tr']/td[6]",
			#'chandi':"",
			};

	def __init__(self):
		WebProcessor_Base.__init__(self,'www.mianzhuang.com')
	
	def GetDetailList(self, batch_num_list):
		try:
			ret_dic = {};
			for batch_num in batch_num_list:
				ret_dic[batch_num] = None;
			for batch_num in batch_num_list:
				page = self.GetDetailPage(batch_num);
				info = self.ParseDetailPage(batch_num, page);
				ret_dic[batch_num]  = info;
				time.sleep(time_per_req)
			return ret_dic;
		except Exception,e:
			raise e;

	def GetInfoByBatchNum(self, batch_num):
		page = self.GetDetailPage(batch_num);
		info = self.ParseDetailPage(batch_num, page);
		return info;
		

	def GetDetailPage( self, sub_url ):
		url = "%s%s"%(self.url_detail_prefix, sub_url);
		req = requests.get(url)
		return req.content

	def ParseDetailPage( self,batch_num, page):
		ret = {};
		html = etree.HTML(page);
		for field in sorted(self.xpath_dic.keys()):
			result = html.xpath(self.xpath_dic[field]);
			if len(result) < 1 :
				#print "%s\t%s less than 1"%(batch_num, field);
				continue;
			if field == 'preparation':
				#pdb.set_trace();
				val = result[0].text.encode('utf-8').strip();
				m = re.match(self.ptn_preparation, val);
				if m:
					ret['preparation_p1'] = m.groupdict()['p1'];
					ret['preparation_p2'] = m.groupdict()['p2'];
					ret['preparation_p3'] = m.groupdict()['p3'];
				else:
					ret['preparation_p1'] = 'unknown';
					ret['preparation_p2'] = 'unknown';
					ret['preparation_p3'] = 'unknown';
			else :
				val = result[0].text.encode('utf-8').strip();
				ret[field] = val;
		ret = self.normaliza_fields(batch_num, ret);
		#for key in ret:
		#	print "%s:%s-dic:%s"%(batch_num,key,ret[key]);
		return ret;


class WebProcessor_EMC(WebProcessor_Base):
	url = '';
	''
	def __init__(self):
		WebProcessor_Base.__init__(self,'www.emiancang.com')

class WebProcessor_CE(WebProcessor_Base):
	url='';
	def __init__(self):
		WebProcessor_Base.__init__(self,'www.cattoneasy.com')


class WebProcessor_DaYuanBo(WebProcessor_Base):
	url = "http://www.dybcotton.com/search/express/json"
	url_detail_prefix = 'http://www.dybcotton.com'
	ptn_processor =  re.compile(r'.*\d+\((?P<name>.*)\)')
	ptn_preparation = re.compile(r'.*P1:(?P<p1>.*)%.*P2:(?P<p2>.*)%.*P3:(?P<p3>.*)%.*')
	ptn_micronaire_avg = re.compile(r'.*平均值:(?P<avg>\S*).*')
	ptn_length_avg = re.compile(r'.*平均值:(?P<avg>\S*).*')

	xpath_dic = {
			'color_11':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[9]/td[3]",
			'color_21':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[10]/td[2]",
			'color_31':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[11]/td[2]",
			'color_41':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[12]/td[2]",
			'color_51':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[13]/td[2]",
			'color_12':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[14]/td[2]",
			'color_22':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[15]/td[2]",
			'color_32':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[16]/td[2]",
			'color_13':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[17]/td[2]",
			'color_23':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[18]/td[2]",
			'color_33':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[19]/td[2]",
			'color_14':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[20]/td[2]",
			'color_24':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[21]/td[2]",
			'length_avg':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[8]/td[4]/text()",
			'length_32':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[9]/td[6]",
			'length_31':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[10]/td[4]",
			'length_30':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[11]/td[4]",
			'length_29':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[12]/td[4]",
			'length_28':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[13]/td[4]",
			'length_27':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[14]/td[4]",
			'length_26':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[15]/td[4]",
			'length_25':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[16]/td[4]",
			'micronaire_avg':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[17]/td[3]/text()",
			'micronaire_c1':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[18]/td[5]",
			'micronaire_b1':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[19]/td[4]",
			'micronaire_a':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[20]/td[4]",
			'micronaire_b2':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[21]/td[4]",
			'micronaire_c2':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[22]/td[4]",
			'breaking_tenacity_avg':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[25]/td[4]",
			'breaking_tenacity_max':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[24]/td[4]",
			'breaking_tenacity_min':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[23]/td[6]",
			'length_uniformty_avg':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[25]/td[2]",
			'length_uniformty_max':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[24]/td[2]",
			'length_uniformty_min':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[23]/td[3]",
			'preparation':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[22]/td[2]",
			'package_num':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[3]/td[2]",
			'weight_gross':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[4]/td[@class='wt_m']",
			'weight_tare':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[5]/td[@class='wt_p']",
			'weight_net':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[6]/td[@class='wt_j']",
			'weight_conditoned':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[6]/td[@class='wt_g']",
			'huichao_avg':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[4]/td[@class='hc']",
			'hanza_avg':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[5]/td[@class='hz']",
			'jiagongleixing':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1 table-striped']/tbody/tr[2]/td[2]",
			'jiagongqiye':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1']/tr[2]/td[6]",
			'cangku':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1']/tr[1]/td[6]",
			'chandi':"/html/body/div[@class='main']/div[@class='main_1']/div[@class='xuqiu']/table[@class='tb1']/tr[2]/td[4]",
			};

	def __init__(self):
		WebProcessor_Base.__init__(self,'www.dybcotton.com')

	def GetDetailList(self, batch_num_list):
		try:
			ret_dic = {};
			for batch_num in batch_num_list:
				ret_dic[batch_num] = None;
			for batch_num in batch_num_list:
				sub_url = self.GetAsinByBatchNum(batch_num);
				time.sleep(time_in_req)
				if None == sub_url or len(sub_url) <= 0:
					#print "error occur in batch_num:%s"%(batch_num);
					continue;
				page = self.GetDetailPage(sub_url);
				info = self.ParseDetailPage(batch_num, page);
				ret_dic[batch_num]  = info;
				time.sleep(time_per_req)
			return ret_dic;
		except Exception,e:
                        print e
			raise e;
	
	def GetInfoByBatchNum(self, batch_num):
		sub_url = self.GetAsinByBatchNum(batch_num);
		if None == sub_url or len(sub_url) <= 0:
			#print "error occur in batch_num:%s"%(batch_num);
			return None;
		time.sleep(time_in_req);
		page = self.GetDetailPage(sub_url);
		info = self.ParseDetailPage(batch_num, page);
		return info;

	def GetAsinByBatchNum(self, batch_num):
		if None == batch_num or len(batch_num) < 1:
			return None;
		params = {'search':batch_num};
		req = requests.post(self.url,data=params,headers=self.headers);
		#pdb.set_trace();
		#TODO exception
		data_dic_list = json.loads(req.content);
		for data_dic in data_dic_list:
			if( 'name' in data_dic and data_dic['name'].find(batch_num) != -1):
				#print "%s\t%s"%(batch_num, data_dic['url']);
				return data_dic['url'];
		return None;

	def GetDetailPage( self,sub_url ):
		url = "%s%s"%(self.url_detail_prefix, sub_url);
		req = requests.get(url)
		return req.content

	def ParseDetailPage( self,batch_num, page):
		ret = {};
		html = etree.HTML(page);
		for field in sorted(self.xpath_dic.keys()):
			result = html.xpath(self.xpath_dic[field]);
			if len(result) < 1 :
				#print "%s\t%s"%(batch_num, field);
				continue;
			if field == 'micronaire_avg':
				ret[field] = 'unknown'
				for txt in result:
					val = txt.encode('utf-8');
					m = re.match(self.ptn_micronaire_avg,val)
					if m:
						ret[field] = m.groupdict()['avg'];
						break;
			elif field == 'preparation':
				val = result[0].text.encode('utf-8').strip();
				m = re.match(self.ptn_preparation, val);
				if m:
					ret['preparation_p1'] = m.groupdict()['p1'];
					ret['preparation_p2'] = m.groupdict()['p2'];
					ret['preparation_p3'] = m.groupdict()['p3'];
				else:
					ret['preparation_p1'] = 'unknown';
					ret['preparation_p2'] = 'unknown';
					ret['preparation_p3'] = 'unknown';
			elif field == 'length_avg':
				ret[field] = 'unknown';
				for txt in result:
					val = txt.encode('utf-8')
					m = re.match(self.ptn_length_avg,val);
					if m:
						ret[field] = m.groupdict()['avg'];
						break;
			elif field == 'jiagongleixing':
				val = result[0].text.encode('utf-8').strip();
				m = re.match(self.ptn_processor, val);
				if m:
					ret[field] = m.groupdict()['name'];
				else:
					ret[field] = 'unknown';
			elif field in ['huichao_avg','hanza_avg']:
                                val = result[0].text.encode('utf-8').strip();
                                if len(val)>1 and val.endswith('%'):
                                        val = val[0:-1]
                                ret[field] = val;
			else :
				val = result[0].text.encode('utf-8').strip();
				ret[field] = val;
		ret = self.normaliza_fields(batch_num, ret);
		#for key in ret:
		#	print "%s:%s-dic:%s"%(batch_num,key,ret[key]);
		return ret;



def thd_call_get_info_list(prod_list, batch_num_list,thd_id=0, thd_num=0):
	#print "%d-%d start"%(thd_num, thd_id);
	try:
		ret_dic = {};
		total_prod_num = len(prod_list);
		for batch_num in batch_num_list:
			if hash(batch_num)% thd_num != thd_id:
				continue;
			ret_dic[batch_num] = None;
		for batch_num in batch_num_list:
			if hash(batch_num)% thd_num != thd_id:
				continue;
			#print "%d-%d\t%s"%(thd_num, thd_id, batch_num);
			prod_id = thd_id if thd_id < total_prod_num else 0;
			prod_num = 1;
			while prod_num < total_prod_num:
				try:
					prod = prod_list[prod_id];
					ret_dic[batch_num]  = prod.GetInfoByBatchNum(batch_num);
					break;
				except Exception,e:
					if prod_num < total_prod_num:
						prod_id = (prod_id+1)%total_prod_num;
						prod_num += 1;
						time.sleep(time_per_req);
					else:
						raise e;
			time.sleep(time_per_req);
		return ret_dic;
	except Exception,e:
		raise e;


class WebProdManager:
	def __init__(self, db_path=None):
		self.parser_dyb = WebProcessor_DaYuanBo();
		self.parser_mz = WebProcessor_MZ();
		self.parser_emc = WebProcessor_EMC();
		self.parser_ce = WebProcessor_CE();
		self.db_path = db_path
		self.db = DataBase(self.db_path);

	
	def GetInfoSingleThread(self, batch_num_list):
		return self.parser_dyb.GetDetailList(batch_num_list);
		#return self.parser_mz.GetDetailList(batch_num_list);
		#return self.parser_emc.GetDetailList(batch_num_list);
		#return self.parser_ce.GetDetailList(batch_num_list);
			
	
	def GetInfoMultiThread(self,batch_num_list):
		parser_list = [self.parser_dyb, self.parser_mz];
		#thd_count = multiprocessing.cpu_count();
		thd_count = len(parser_list);
		pool = multiprocessing.Pool( thd_count ) ;
		dic_merge = {};
		result = [];
		try:
			for thread_id in range(thd_count):
				res = pool.apply_async(thd_call_get_info_list, args=(parser_list,batch_num_list,thread_id,thd_count,));
				result.append( res );
			pool.close();
			pool.join();
			for res in result:
				dic = res.get();
				dic_merge = dict(dic_merge, **dic);
			return dic_merge;
		except KeyboardInterrupt:
			pool.terminal();
			pool.join();

	def is_detail_info_available(self, src_dic):
		if None == src_dic:
			return False;
		num_hits_db_column_name = 0;
		num_empty = 0;
		num_column = len(self.db.catton_column_info);
		for key in self.db.catton_column_info:
			if key not in src_dic:
				continue;
			num_hits_db_column_name += 1;
			v = src_dic[key];
			if None == v or ( not isinstance(v,str) and not isinstance(v, unicode) )or len(v) < 1:
				num_empty += 1;
		#print num_column,num_hits_db_column_name,num_empty
		if num_hits_db_column_name < num_column or num_column - num_empty < 10:
			return False;
		
		return True;

	def GetWebData(self,batch_num_list):
		batch_num_list_todo = [];
		ret_dic = {};
		for batch_num in batch_num_list:
			db_ret = self.db.get_cotton_by_batch_num(batch_num);
			#print batch_num,self.is_detail_info_available(db_ret);
			#print db_ret;
			if not self.is_detail_info_available(db_ret):
				batch_num_list_todo.append(batch_num);
			else:
				ret_dic[batch_num] = db_ret;
		
		web_ret = {};
		if 0 == len(batch_num_list_todo):
			pass;
		elif len(batch_num_list_todo) < 10:
			web_ret = self.GetInfoSingleThread(batch_num_list_todo);
		else:
			web_ret = self.GetInfoMultiThread(batch_num_list_todo);

		for k in web_ret.keys():
			if not self.is_detail_info_available(web_ret[k]):
				continue;
			self.db.add_cotton(web_ret[k]);
			#print k,self.is_detail_info_available(web_ret[k]);
			#print web_ret[k];
		ret_dic = dict(ret_dic,**web_ret);
		return ret_dic;



class ExcelTemplete:

	def __init__(self,ptn_id):
		if 1 == ptn_id:
			self.get_premium_discount_pattern();
		else:
			self.get_full_pattern();
	
	def get_premium_discount_pattern(self):
                self.head_row = 0;
                self.content_row = 1;
		self.head_list = [
				[u'批号' , [0,0]],
				[u'11' , [0,1]],
				[u'21' , [0,2]],
				[u'31' , [0,3]],
				[u'41' , [0,4]],
				[u'51' , [0,5]],
				[u'12' , [0,6]],
				[u'22' , [0,7]],
				[u'32' , [0,8]],
				[u'13' , [0,9]],
				[u'23' , [0,10]],
				[u'33' , [0,11]],
				[u'14' , [0,12]],
				[u'24' , [0,13]],
				[u'平均长度' , [0,14]],
				[u'平均马值' , [0,15]],
				[u'平均断裂比强度' , [0,16]],
				[u'平均长度整齐度' , [0,17]],
				[u'P1' , [0,18]],
				[u'P2' , [0,19]],
				[u'P3' , [0,20]],
				];
		self.dic_pos = {
				'batch_num':0,
				'color_11':1,
				'color_21':2,
				'color_31':3,
				'color_41':4,
				'color_51':5,
				'color_12':6,
				'color_22':7,
				'color_32':8,
				'color_13':9,
				'color_23':10,
				'color_33':11,
				'color_14':12,
				'color_24':13,
				'length_avg':14,
				'micronaire_avg':15,
				'breaking_tenacity_avg':16,
				'length_uniformty_avg':17,
				'preparation_p1':18,
				'preparation_p2':19,
				'preparation_p3':20,
				};

	def get_full_pattern(self):
                self.head_row = 0;
                self.content_row = 2;
		self.head_list = [
				[u'批号' , [0,1,0,0]],
				[u'颜色级' , [0,0,1,13]],
				[u'11' , [1,1]],
				[u'21' , [1,2]],
				[u'31' , [1,3]],
				[u'41' , [1,4]],
				[u'51' , [1,5]],
				[u'12' , [1,6]],
				[u'22' , [1,7]],
				[u'32' , [1,8]],
				[u'13' , [1,9]],
				[u'23' , [1,10]],
				[u'33' , [1,11]],
				[u'14' , [1,12]],
				[u'24' , [1,13]],
				[u'平均长度' , [0,1,14,14]],
				[u'长度分布' , [0,0,15,22]],
				[u'32mm' , [1,15]],
				[u'31mm' , [1,16]],
				[u'30mm' , [1,17]],
				[u'29mm' , [1,18]],
				[u'28mm' , [1,19]],
				[u'27mm' , [1,20]],
				[u'26mm' , [1,21]],
				[u'25mm' , [1,22]],
				[u'马克隆平均值' , [0,1,23,23]],
				[u'马值分布' , [0,0,24,28]],
				[u'C1' , [1,24]],
				[u'B1' , [1,25]],
				[u'A' , [1,26]],
				[u'B2' , [1,27]],
				[u'C2' , [1,28]],
				[u'断裂比强度' , [0,0,29,31]],
				[u'平均值' , [1,29]],
				[u'最大值' , [1,30]],
				[u'最小值' , [1,31]],
				[u'长度整齐度' , [0,0,32,34]],
				[u'平均值' , [1,32]],
				[u'最大值' , [1,33]],
				[u'最小值' , [1,34]],
				[u'轧工质量' , [0,0,35,37]],
				[u'P1' , [1,35]],
				[u'P2' , [1,36]],
				[u'P3' , [1,37]],
				[u'合计包数' , [0,1,38,38]],
				[u'合计毛重' , [0,1,39,39]],
				[u'合计皮重' , [0,1,40,40]],
				[u'合计净重' , [0,1,41,41]],
				[u'合计公重' , [0,1,42,42]],
				[u'平均回潮' , [0,1,43,43]],
				[u'平均含杂' , [0,1,44,44]],
				[u'产地' , [0,1,45,45]],
				[u'加工类型' , [0,1,46,46]],
				[u'加工企业' , [0,1,47,47]],
				[u'仓库' , [0,1,48,48]],
				];
		self.dic_pos = {
				'batch_num':0,
				'color_11':1,
				'color_21':2,
				'color_31':3,
				'color_41':4,
				'color_51':5,
				'color_12':6,
				'color_22':7,
				'color_32':8,
				'color_13':9,
				'color_23':10,
				'color_33':11,
				'color_14':12,
				'color_24':13,
				'length_avg':14,
				'length_32':15,
				'length_31':16,
				'length_30':17,
				'length_29':18,
				'length_28':19,
				'length_27':20,
				'length_26':21,
				'length_25':22,
				'micronaire_avg':23,
				'micronaire_c1':24,
				'micronaire_b1':25,
				'micronaire_a':26,
				'micronaire_b2':27,
				'micronaire_c2':28,
				'breaking_tenacity_avg':29,
				'breaking_tenacity_max':30,
				'breaking_tenacity_min':31,
				'length_uniformty_avg':32,
				'length_uniformty_max':33,
				'length_uniformty_min':34,
				'preparation_p1':35,
				'preparation_p2':36,
				'preparation_p3':37,
				'package_num':38,
				'weight_gross':39,
				'weight_tare':40,
				'weight_net':41,
				'weight_conditoned':42,
				'huichao_avg':43,
				'hanza_avg':44,
				'jiagongleixing':46,
				'jiagongqiye':47,
				'cangku':48,
				'chandi':45,
				};


class FileExcel_XLSX:
	def __init__(self):
		self.ptn_batch_num = re.compile(r'^[0-9]+$');
	
	def writeHead(self, table, start_row, template):
                start_row = template.head_row;
		head_list = template.head_list;
		for item in head_list:
			if len(item) != 2:
				continue;
			#label = item[0].encode('utf-8');
			label = item[0];
			pos = item[1];
			if 2 == len(pos):
				table.cell(row=start_row+pos[0]+1,column=pos[1]+1, value=label);
			elif 4 == len(pos):
				table.merge_cells(start_row=start_row+pos[0]+1,start_column=pos[2]+1,end_row=start_row + pos[1]+1,end_column=pos[3]+1);
				table.cell(row=start_row+pos[0]+1,column=pos[2]+1, value=label);
			else:
				continue;

	def writeContent(self, table, start_row, dic_val, template, order_list):
                start_row = template.content_row;
		dic_pos = template.dic_pos;
		row_num = start_row-1;
		for idx in range(len(order_list)):
			batch_num = order_list[idx];
			if None == batch_num or batch_num not in dic_val:
				continue;
			row_num += 1;
			if 'batch_num' in dic_pos:
				col = dic_pos['batch_num'];
				table.cell(row=row_num+1,column=col+1,value=batch_num);
			info = dic_val[batch_num];
			if None == info:
				continue;
			for seg in info.keys():
				if seg not in dic_pos:
					continue;
				col = dic_pos[seg];
				table.cell(row=row_num+1,column=col+1,value=info[seg]);

	def SaveFile(self, path, dic_val, template, order_list):
		wb=openpyxl.Workbook();
		table = wb.active;
		self.writeHead(table,0,template);
		self.writeContent(table,2,dic_val,template,order_list);
		wb.save(path);

	def getBatchNumListFromExcel(self,filePath, columnId):
		try:
			wb=load_workbook(filePath);
		except Exception,e:
			raise e

class FileExcel_XLS:
	def __init__(self):
		self.ptn_batch_num = re.compile(r'^[0-9]+$');
		self.style_head = xlwt.XFStyle();
		self.style_content = xlwt.XFStyle();
		
	
	def writeHead(self, table, start_row, template):
                start_row = template.head_row;
		head_list = template.head_list;
		for item in head_list:
			if len(item) != 2:
				continue;
			#label = item[0].encode('utf-8');
			label = item[0];
			pos = item[1];
			if 2 == len(pos):
				table.write(start_row+pos[0], pos[1], label);
			elif 4 == len(pos):
				table.write_merge(start_row+pos[0],start_row + pos[1],pos[2],pos[3],label);
			else:
				continue;

	def writeContent(self, table, start_row, dic_val, template, order_list):
                start_row = template.content_row;
		dic_pos = template.dic_pos;
		row_num = start_row-1;
		for idx in range(len(order_list)):
			batch_num = order_list[idx];
			if None == batch_num or batch_num not in dic_val:
				continue;
			row_num += 1;
			if 'batch_num' in dic_pos:
				col = dic_pos['batch_num'];
				table.write(row_num, col, batch_num);
			info = dic_val[batch_num];
			if None == info:
				continue;
			for seg in info.keys():
				if seg not in dic_pos:
					continue;
				col = dic_pos[seg];
				table.write(row_num,col,info[seg]);

	def SaveFile(self, path, dic_val, template, order_list):
		wb = xlwt.Workbook(encoding='utf-8');
		table = wb.add_sheet(u'sheet1',cell_overwrite_ok=True);
		self.writeHead(table,0,template);
		self.writeContent(table,2,dic_val,template, order_list);
		wb.save(path);

	def getBatchNumListFromExcel(self,filePath, columnId):
		try:
			ret_list = [];
			data = xlrd.open_workbook(filePath);
			table = data.sheet_by_index(0);
			for i in range(table.nrows):
				val = table.cell(i, columnId).value;
				if type(val) is types.IntType:
					ret_list.append(str(val));
				elif type(val) is types.FloatType:
					ret_list.append(str("%.0f"%(val)));
				elif type(val) is types.StringType:
					if re.match(self.ptn_batch_num,val):
						ret_list.append(val);
				elif type(val) is types.UnicodeType:
					val = val.encode('utf-8');
					if re.match(self.ptn_batch_num,val):
						ret_list.append(val);
			return ret_list;
		except Exception,e:
			raise e



class FileManager:
	def __init__(self):
		self.file_excel_xls = FileExcel_XLS();
		self.file_excel_xlsx = FileExcel_XLSX();

	def ReadBatchNumList(self,file_path,column_id, file_type='excel'):
		if file_type == 'excel':
			return self.file_excel_xls.getBatchNumListFromExcel(file_path, column_id);

	def save(self, path, dic_val, template_id, order_list, file_type='excel'):
		template = ExcelTemplete(template_id);
		if file_type == 'excel':
			if path.endswith('xlsx'):
				self.file_excel_xlsx.SaveFile(path,dic_val, template, order_list);
			else:
				self.file_excel_xls.SaveFile(path,dic_val, template, order_list);


if __name__ == '__main__':
	#batch_num_list = ['65140161194','65140161191'];
	page_mgr = WebProdManager();
	file_mgr = FileManager();
	batch_num_list = file_mgr.ReadBatchNumList('/Users/jiangnz/workspace/cis/a.xlsx',0);
	dic_info = page_mgr.GetWebData(batch_num_list);
	file_mgr.save('/Users/jiangnz/workspace/cis/demo.xls', dic_info);
	#if None != dic_info:
	#	for key in dic_info.keys():
	#		for id in dic_info[key]:
	#			print "%s\t%s\t%s"%(key, id, dic_info[key][id]);

